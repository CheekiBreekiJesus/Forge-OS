"""Reusable lead database preparation pipeline.

The module is intentionally generic: customer-specific paths and optional
mapping files are provided at runtime and must stay outside Git.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from lead_schema import (
    FIELD_ALIASES,
    FIELD_DEFINITIONS,
    LOOKUPS,
    SOURCE_COLUMNS,
    STANDARD_FIELDS,
)

SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".tsv"}
SCRIPT_VERSION = "2026.06.29"


@dataclass
class SourceInfo:
    source_id: str
    source_file: str
    source_path: str
    source_sheet: str
    source_database: str
    source_tab_name: str
    extension: str
    row_count: int
    column_count: int
    header_row: int | None
    columns: list[str]
    sample_values: dict[str, list[str]]
    merged_cells: list[str] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_columns: list[str] = field(default_factory=list)
    formula_cells: int = 0
    notes: list[str] = field(default_factory=list)
    imported_lead_count: int = 0
    duplicate_count: int = 0
    error_count: int = 0
    excluded_count: int = 0


@dataclass
class PreparedData:
    source_infos: list[SourceInfo]
    source_rows: list[dict[str, Any]]
    master_rows: list[dict[str, Any]]
    field_mappings: list[dict[str, Any]]
    duplicates: list[dict[str, Any]]
    import_errors: list[dict[str, Any]]
    excluded_rows: list[dict[str, Any]]
    change_log: list[dict[str, Any]]


def discover_sources(input_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in input_dir.rglob("*")
        if path.is_file()
        and path.suffix.lower() in SUPPORTED_EXTENSIONS
        and not path.name.startswith("~$")
        and "standardized" not in path.name.lower()
    )


def inspect_sources(input_dir: Path) -> list[SourceInfo]:
    source_infos: list[SourceInfo] = []
    for file_index, path in enumerate(discover_sources(input_dir), start=1):
        if path.suffix.lower() == ".xlsx":
            source_infos.extend(_inspect_xlsx(path, file_index))
        else:
            source_infos.append(_inspect_csv(path, file_index))
    return source_infos


def prepare_data(input_dir: Path, import_batch: str | None = None) -> PreparedData:
    import_batch = import_batch or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    source_infos = inspect_sources(input_dir)
    all_source_rows: list[dict[str, Any]] = []
    field_mappings: list[dict[str, Any]] = []
    import_errors: list[dict[str, Any]] = []
    excluded_rows: list[dict[str, Any]] = []
    change_log: list[dict[str, Any]] = []

    for info in source_infos:
        raw_rows = _read_source_rows(info)
        mappings = [_build_field_mapping(info, column, raw_rows) for column in info.columns]
        field_mappings.extend(mappings)
        mapping_by_column = {mapping["original_column"]: mapping["standardized_column"] for mapping in mappings}

        for raw_row in raw_rows:
            standardized, error = _standardize_row(raw_row, info, mapping_by_column, import_batch)
            source_record = {
                **{key: standardized.get(key, "") for key in SOURCE_COLUMNS if key != "raw_source_data"},
                "raw_source_data": standardized.get("raw_source_data", ""),
                "standardized": standardized,
            }

            if error:
                import_errors.append(error)
                info.error_count += 1
            elif _is_potential_lead(standardized):
                info.imported_lead_count += 1
                all_source_rows.append(source_record)
            else:
                info.excluded_count += 1
                excluded_rows.append(
                    {
                        "source_file": info.source_file,
                        "source_sheet": info.source_sheet,
                        "source_row": raw_row["__source_row"],
                        "raw_source_data": standardized["raw_source_data"],
                        "error_type": "excluded_non_lead_row",
                        "error_description": "Row did not contain enough lead signals.",
                        "recommended_action": "Review only if this source should be imported as leads.",
                    }
                )

    master_rows, duplicates = deduplicate_source_rows(all_source_rows)
    duplicate_source_ids = {duplicate["related_lead_id"] for duplicate in duplicates}
    for source_row in all_source_rows:
        source_row["duplicate_group_id"] = next(
            (
                duplicate["duplicate_group_id"]
                for duplicate in duplicates
                if duplicate["related_lead_id"] == source_row["standardized"]["lead_id"]
                or duplicate["canonical_lead_id"] == source_row["standardized"]["lead_id"]
            ),
            "",
        )
        source_row["duplicate_status"] = (
            "exact_duplicate" if source_row["standardized"]["lead_id"] in duplicate_source_ids else source_row["standardized"]["duplicate_status"]
        )
        source_row["standardized"]["duplicate_status"] = source_row["duplicate_status"]

    duplicate_counts = Counter(row["source_file"] for row in all_source_rows if row["duplicate_status"] == "exact_duplicate")
    for info in source_infos:
        info.duplicate_count = duplicate_counts[info.source_file]

    change_log.append(
        {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "transformation": "lead_standardization_pipeline",
            "affected_source": str(input_dir),
            "affected_rows": len(all_source_rows),
            "description": "Inspected, standardized, validated, and deduplicated supported lead source files.",
            "script_version": SCRIPT_VERSION,
        }
    )

    return PreparedData(
        source_infos=source_infos,
        source_rows=all_source_rows,
        master_rows=master_rows,
        field_mappings=field_mappings,
        duplicates=duplicates,
        import_errors=import_errors,
        excluded_rows=excluded_rows,
        change_log=change_log,
    )


def deduplicate_source_rows(source_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    canonical_by_key: dict[str, dict[str, Any]] = {}
    possible_name_groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    master_rows: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []

    for source_row in source_rows:
        row = dict(source_row["standardized"])
        exact_key = _exact_duplicate_key(row)
        if exact_key and exact_key in canonical_by_key:
            canonical = canonical_by_key[exact_key]
            duplicate_group_id = f"DUP-{len(duplicates) + 1:05d}"
            _merge_non_conflicting(canonical, row)
            duplicates.append(
                {
                    "duplicate_group_id": duplicate_group_id,
                    "duplicate_type": "exact",
                    "confidence": "high",
                    "canonical_lead_id": canonical["lead_id"],
                    "related_lead_id": row["lead_id"],
                    "source_file": row["source_file"],
                    "source_row": row["source_row"],
                    "matching_fields": exact_key,
                    "conflicting_fields": "",
                    "recommended_action": "Merged into canonical record; source reference preserved.",
                    "review_status": "merged",
                }
            )
            canonical["duplicate_status"] = "exact_duplicate"
            canonical["source_database"] = _append_unique(canonical["source_database"], row["source_database"])
            canonical["source_file"] = _append_unique(canonical["source_file"], row["source_file"])
            canonical["source_sheet"] = _append_unique(canonical["source_sheet"], row["source_sheet"])
            canonical["source_row"] = _append_unique(str(canonical["source_row"]), str(row["source_row"]))
            continue

        if exact_key:
            canonical_by_key[exact_key] = row

        name_key = _possible_duplicate_key(row)
        if name_key:
            possible_name_groups[name_key].append(row)

        master_rows.append(row)

    for group_rows in possible_name_groups.values():
        if len(group_rows) < 2:
            continue
        duplicate_group_id = f"POSS-{len(duplicates) + 1:05d}"
        canonical = group_rows[0]
        for related in group_rows[1:]:
            if related["lead_id"] == canonical["lead_id"]:
                continue
            related["duplicate_status"] = "possible_duplicate"
            related["review_required"] = "true"
            duplicates.append(
                {
                    "duplicate_group_id": duplicate_group_id,
                    "duplicate_type": "possible",
                    "confidence": "medium",
                    "canonical_lead_id": canonical["lead_id"],
                    "related_lead_id": related["lead_id"],
                    "source_file": related["source_file"],
                    "source_row": related["source_row"],
                    "matching_fields": "normalized company plus city/domain signal",
                    "conflicting_fields": "",
                    "recommended_action": "Manual review before merging.",
                    "review_status": "open",
                }
            )

    for index, row in enumerate(master_rows, start=1):
        row["lead_id"] = f"LEAD-{index:06d}"
        row["company_id"] = row.get("company_id") or _stable_id("COMP", row.get("company_name_normalized") or row["lead_id"])
        row["contact_id"] = row.get("contact_id") or _stable_id("CONT", (row.get("email") or row.get("contact_name") or row["lead_id"]))

    return master_rows, duplicates


def build_workbook(prepared: PreparedData, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary(wb, prepared)
    _write_table_sheet(wb, "MASTER_LEADS", STANDARD_FIELDS, prepared.master_rows, highlight_review=True)
    _write_source_index(wb, prepared)

    for info in prepared.source_infos:
        rows = [
            _flatten_source_row(row)
            for row in prepared.source_rows
            if row["standardized"]["source_file"] == info.source_file and row["standardized"]["source_sheet"] == info.source_sheet
        ]
        rows.extend(
            {
                "source_file": excluded["source_file"],
                "source_sheet": excluded["source_sheet"],
                "source_row": excluded["source_row"],
                "source_database": info.source_database,
                "standardized_lead_id": "",
                "duplicate_group_id": "",
                "duplicate_status": "excluded",
                "validation_status": "excluded",
                "review_required": "true",
                "raw_source_data": excluded["raw_source_data"],
            }
            for excluded in prepared.excluded_rows
            if excluded["source_file"] == info.source_file and excluded["source_sheet"] == info.source_sheet
        )
        rows.extend(
            {
                "source_file": error["source_file"],
                "source_sheet": error["source_sheet"],
                "source_row": error["source_row"],
                "source_database": info.source_database,
                "standardized_lead_id": "",
                "duplicate_group_id": "",
                "duplicate_status": "error",
                "validation_status": "error",
                "review_required": "true",
                "raw_source_data": error["raw_source_data"],
            }
            for error in prepared.import_errors
            if error["source_file"] == info.source_file and error["source_sheet"] == info.source_sheet
        )
        _write_table_sheet(wb, info.source_tab_name, SOURCE_COLUMNS, rows)

    _write_table_sheet(wb, "FIELD_MAPPING", list(prepared.field_mappings[0].keys()) if prepared.field_mappings else [], prepared.field_mappings)
    _write_table_sheet(wb, "DUPLICATES", list(prepared.duplicates[0].keys()) if prepared.duplicates else _duplicate_headers(), prepared.duplicates)
    _write_table_sheet(wb, "IMPORT_ERRORS", _import_error_headers(), prepared.import_errors + prepared.excluded_rows)
    _write_lookups(wb, prepared)
    _write_data_dictionary(wb)
    _write_table_sheet(wb, "CHANGE_LOG", list(prepared.change_log[0].keys()), prepared.change_log)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def validate_workbook(output_path: Path, expected_source_rows: int | None = None) -> dict[str, Any]:
    wb = load_workbook(output_path, data_only=False)
    required_sheets = {
        "SUMMARY",
        "MASTER_LEADS",
        "SOURCE_INDEX",
        "FIELD_MAPPING",
        "DUPLICATES",
        "IMPORT_ERRORS",
        "LOOKUPS",
        "DATA_DICTIONARY",
        "CHANGE_LOG",
    }
    missing_sheets = sorted(required_sheets - set(wb.sheetnames))
    master = wb["MASTER_LEADS"]
    headers = [cell.value for cell in master[1]]
    lead_id_col = headers.index("lead_id") + 1
    source_file_col = headers.index("source_file") + 1
    lead_ids = [master.cell(row=row, column=lead_id_col).value for row in range(2, master.max_row + 1)]
    source_refs = [master.cell(row=row, column=source_file_col).value for row in range(2, master.max_row + 1)]
    formula_errors = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
    source_sheet_rows = sum(
        max(0, sheet.max_row - 1)
        for sheet in wb.worksheets
        if sheet.title.startswith("SOURCE_") and sheet.title != "SOURCE_INDEX"
    )
    source_row_count_matches = expected_source_rows is None or source_sheet_rows == expected_source_rows

    result = {
        "valid": not missing_sheets
        and len(lead_ids) == len(set(lead_ids))
        and all(source_refs)
        and not formula_errors
        and "Sheet" not in wb.sheetnames
        and source_row_count_matches,
        "missing_sheets": missing_sheets,
        "master_leads": len(lead_ids),
        "unique_lead_ids": len(set(lead_ids)),
        "records_with_source_refs": sum(1 for ref in source_refs if ref),
        "formula_errors": formula_errors,
        "expected_source_rows": expected_source_rows,
        "source_sheet_rows": source_sheet_rows,
        "source_row_count_matches": source_row_count_matches,
        "workbook_sheets": wb.sheetnames,
    }
    return result


def write_json_report(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_inspection_report(source_infos: list[SourceInfo], path: Path) -> None:
    lines = ["# Lead Source Inspection Report", ""]
    for info in source_infos:
        lines.extend(
            [
                f"## {info.source_file} / {info.source_sheet}",
                "",
                f"- Source database: {info.source_database}",
                f"- Rows: {info.row_count}",
                f"- Columns: {info.column_count}",
                f"- Header row: {info.header_row}",
                f"- Generated source tab: {info.source_tab_name}",
                f"- Formula cells: {info.formula_cells}",
                f"- Merged cells: {len(info.merged_cells)}",
                f"- Hidden rows: {len(info.hidden_rows)}",
                f"- Hidden columns: {len(info.hidden_columns)}",
                f"- Columns: {', '.join(info.columns)}",
                "",
            ]
        )
        if info.notes:
            lines.append("- Notes:")
            lines.extend(f"  - {note}" for note in info.notes)
            lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def run_pipeline(input_dir: Path, output_dir: Path, workbook_name: str = "Lead_Database_Standardized.xlsx") -> dict[str, Any]:
    source_paths = discover_sources(input_dir)
    mtimes_before = {str(path): path.stat().st_mtime_ns for path in source_paths}
    prepared = prepare_data(input_dir)
    workbook_path = output_dir / workbook_name
    report_path = output_dir / "inspection_report.md"
    validation_path = output_dir / "validation_report.json"
    summary_path = output_dir / "pipeline_summary.json"

    build_workbook(prepared, workbook_path)
    validation = validate_workbook(workbook_path, expected_source_rows=len(prepared.source_rows) + len(prepared.excluded_rows) + len(prepared.import_errors))
    mtimes_after = {str(path): path.stat().st_mtime_ns for path in source_paths}
    validation["source_files_unchanged"] = mtimes_before == mtimes_after
    validation["valid"] = validation["valid"] and validation["source_files_unchanged"]

    write_inspection_report(prepared.source_infos, report_path)
    write_json_report(validation_path, validation)
    write_json_report(summary_path, _summary_payload(prepared, workbook_path, report_path, validation_path, validation))

    return {
        "prepared": prepared,
        "workbook_path": workbook_path,
        "inspection_report_path": report_path,
        "validation_report_path": validation_path,
        "summary_path": summary_path,
        "validation": validation,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Prepare standardized lead workbook.")
    parser.add_argument("--input", required=True, type=Path, help="Directory containing source Excel/CSV files.")
    parser.add_argument("--output", required=True, type=Path, help="Directory for generated local outputs.")
    parser.add_argument("--workbook-name", default="Lead_Database_Standardized.xlsx")
    args = parser.parse_args(argv)
    result = run_pipeline(args.input, args.output, args.workbook_name)
    print(json.dumps(_printable_result(result), ensure_ascii=False, indent=2))
    return 0 if result["validation"]["valid"] else 1


def _inspect_xlsx(path: Path, file_index: int) -> list[SourceInfo]:
    wb = load_workbook(path, data_only=False, read_only=False)
    infos = []
    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        values = list(ws.iter_rows(values_only=True))
        header_row, columns = _detect_header(values)
        samples = _sample_values(values, header_row, columns)
        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
        hidden_columns = [col for col, dim in ws.column_dimensions.items() if dim.hidden]
        formula_cells = sum(
            1
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        infos.append(
            SourceInfo(
                source_id=f"SRC-{file_index:02d}-{sheet_index:02d}",
                source_file=path.name,
                source_path=str(path),
                source_sheet=ws.title,
                source_database=_source_database_label(path),
                source_tab_name=f"SOURCE_{len(infos) + file_index:02d}_{_safe_sheet_suffix(ws.title)}"[:31],
                extension=path.suffix.lower(),
                row_count=ws.max_row,
                column_count=ws.max_column,
                header_row=header_row,
                columns=columns,
                sample_values=samples,
                merged_cells=[str(rng) for rng in ws.merged_cells.ranges],
                hidden_rows=hidden_rows,
                hidden_columns=hidden_columns,
                formula_cells=formula_cells,
                notes=_inspection_notes(values, header_row, columns),
            )
        )
    return infos


def _inspect_csv(path: Path, file_index: int) -> SourceInfo:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))
    header_row, columns = _detect_header([tuple(row) for row in rows])
    return SourceInfo(
        source_id=f"SRC-{file_index:02d}-01",
        source_file=path.name,
        source_path=str(path),
        source_sheet=path.stem,
        source_database=_source_database_label(path),
        source_tab_name=f"SOURCE_{file_index:02d}_{_safe_sheet_suffix(path.stem)}"[:31],
        extension=path.suffix.lower(),
        row_count=len(rows),
        column_count=max((len(row) for row in rows), default=0),
        header_row=header_row,
        columns=columns,
        sample_values=_sample_values([tuple(row) for row in rows], header_row, columns),
        notes=_inspection_notes([tuple(row) for row in rows], header_row, columns),
    )


def _read_source_rows(info: SourceInfo) -> list[dict[str, Any]]:
    path = Path(info.source_path)
    if info.extension == ".xlsx":
        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb[info.source_sheet]
        values = list(ws.iter_rows(values_only=True))
    else:
        delimiter = "\t" if info.extension == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            values = [tuple(row) for row in csv.reader(handle, delimiter=delimiter)]

    if not info.header_row:
        return []

    data_rows = []
    for row_index, row in enumerate(values[info.header_row :], start=info.header_row + 1):
        if not any(_clean_cell(value) for value in row):
            continue
        raw = {column: _clean_cell(row[idx]) if idx < len(row) else "" for idx, column in enumerate(info.columns)}
        raw["__source_row"] = row_index
        data_rows.append(raw)
    return data_rows


def _standardize_row(raw_row: dict[str, Any], info: SourceInfo, mapping_by_column: dict[str, str], import_batch: str) -> tuple[dict[str, Any], dict[str, Any] | None]:
    row = {field: "" for field in STANDARD_FIELDS}
    row.update(
        {
            "source_file": info.source_file,
            "source_sheet": info.source_sheet,
            "source_row": raw_row["__source_row"],
            "source_database": info.source_database,
            "import_batch": import_batch,
            "lead_status": "new",
            "lead_priority": "medium",
            "validation_status": "partial",
            "duplicate_status": "unique",
            "enrichment_status": "not_started",
            "review_required": "false",
            "consent_status": "unknown",
            "do_not_contact": "false",
            "unsubscribe_status": "unknown",
            "campaign_status": "not_started",
            "country": "",
            "created_at": import_batch,
            "updated_at": import_batch,
        }
    )
    original = {key: value for key, value in raw_row.items() if not key.startswith("__")}
    row["raw_source_data"] = json.dumps(original, ensure_ascii=False, sort_keys=True)

    for original_column, value in original.items():
        target = mapping_by_column.get(original_column, "")
        if not target or not value:
            continue
        if target in row and not row[target]:
            row[target] = value
        elif target in {"email", "phone", "mobile_phone", "website"}:
            _append_secondary(row, target, value)
        else:
            row["notes"] = _append_unique(row["notes"], f"{original_column}: {value}")

    row["raw_company_name"] = row["company_name"]
    row["raw_contact_name"] = row["contact_name"]
    row["raw_email"] = row["email"]
    row["raw_phone"] = row["phone"] or row["mobile_phone"]
    row["raw_address"] = row["address"]
    row["raw_notes"] = row["notes"]

    row["company_name"] = _normalize_spaces(row["company_name"])
    row["company_name_normalized"] = normalize_company_name(row["company_name"])
    row["email"], row["secondary_email"], row["email_status"] = normalize_emails(row["email"])
    row["phone"], row["secondary_phone"], row["phone_status"] = normalize_phones(row["phone"] or row["mobile_phone"])
    row["website"], row["website_status"] = normalize_website(row["website"])
    row["review_required"] = "true" if row["email_status"] == "invalid" or row["phone_status"] == "invalid" else row["review_required"]
    row["validation_status"] = _validation_status(row)
    row["data_quality_score"] = str(_data_quality_score(row))
    row["lead_id"] = _stable_id("RAW", f"{info.source_file}|{info.source_sheet}|{raw_row['__source_row']}")
    row["original_record_id"] = row["lead_id"]

    if not any([row["company_name"], row["contact_name"], row["email"], row["phone"], row["website"]]):
        return row, {
            "source_file": info.source_file,
            "source_sheet": info.source_sheet,
            "source_row": raw_row["__source_row"],
            "raw_source_data": row["raw_source_data"],
            "error_type": "empty_or_unusable_row",
            "error_description": "No lead-identifying fields were detected.",
            "recommended_action": "Review source structure or mapping.",
        }

    return row, None


def _build_field_mapping(info: SourceInfo, column: str, raw_rows: list[dict[str, Any]]) -> dict[str, Any]:
    normalized = _normalize_header(column)
    mapped = FIELD_ALIASES.get(normalized, "")
    samples = []
    for row in raw_rows:
        value = _clean_cell(row.get(column, ""))
        if value and value not in samples:
            samples.append(value)
        if len(samples) >= 3:
            break
    return {
        "source_file": info.source_file,
        "source_sheet": info.source_sheet,
        "original_column": column,
        "sample_values": " | ".join(samples),
        "standardized_column": mapped,
        "transformation_rule": "alias_map_direct" if mapped else "preserved_in_raw_source_data",
        "confidence": "high" if mapped else "low",
        "notes": "" if mapped else "No confident standard field; value preserved in raw_source_data.",
    }


def normalize_company_name(value: str) -> str:
    value = _normalize_spaces(value).casefold()
    value = re.sub(r"[.,;:/\\|_\\-]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_emails(value: str) -> tuple[str, str, str]:
    if not value:
        return "", "", "missing"
    parts = _split_multi_value(value)
    cleaned = [re.sub(r"\s+", "", part).lower().replace("..", ".") for part in parts if part]
    valid = [email for email in cleaned if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email)]
    invalid = [email for email in cleaned if email not in valid]
    if invalid and not valid:
        return cleaned[0], "; ".join(cleaned[1:]), "invalid"
    return valid[0] if valid else "", "; ".join(valid[1:] + invalid), "valid" if not invalid else "review_required"


def normalize_phones(value: str) -> tuple[str, str, str]:
    if not value:
        return "", "", "missing"
    parts = _split_multi_value(value)
    normalized = []
    invalid = []
    for part in parts:
        cleaned = re.sub(r"[^\d+]", "", part)
        digits = re.sub(r"\D", "", cleaned)
        if 9 <= len(digits) <= 15:
            normalized.append(cleaned)
        else:
            invalid.append(part)
    if invalid and not normalized:
        return value, "", "invalid"
    return normalized[0] if normalized else "", "; ".join(normalized[1:] + invalid), "valid" if not invalid else "review_required"


def normalize_website(value: str) -> tuple[str, str]:
    if not value:
        return "", "missing"
    first = _split_multi_value(value)[0].strip()
    if not first:
        return "", "missing"
    if "://" not in first:
        first = f"https://{first}"
    parsed = urlsplit(first)
    if not parsed.netloc:
        return value, "invalid"
    query = urlencode([(key, val) for key, val in parse_qsl(parsed.query) if not key.lower().startswith("utm_")])
    normalized = urlunsplit((parsed.scheme.lower(), parsed.netloc.lower(), parsed.path.rstrip("/"), query, ""))
    return normalized, "valid"


def _detect_header(values: list[tuple[Any, ...]]) -> tuple[int | None, list[str]]:
    best_index = None
    best_score = -1
    for index, row in enumerate(values[:30], start=1):
        cells = [_clean_cell(value) for value in row]
        nonblank = [cell for cell in cells if cell]
        if len(nonblank) < 2:
            continue
        alias_hits = sum(1 for cell in nonblank if _normalize_header(cell) in FIELD_ALIASES)
        text_hits = sum(1 for cell in nonblank if re.search(r"email|telefone|contact|empresa|nome|site|morada", cell, re.I))
        score = alias_hits * 3 + text_hits + min(len(nonblank), 8)
        if score > best_score:
            best_score = score
            best_index = index
    if best_index is None:
        return None, []
    raw_headers = [_clean_cell(value) for value in values[best_index - 1]]
    headers = []
    seen: Counter[str] = Counter()
    for idx, header in enumerate(raw_headers, start=1):
        header = header or f"column_{idx}"
        seen[header] += 1
        headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")
    return best_index, headers


def _sample_values(values: list[tuple[Any, ...]], header_row: int | None, columns: list[str]) -> dict[str, list[str]]:
    if not header_row:
        return {}
    samples: dict[str, list[str]] = {column: [] for column in columns}
    for row in values[header_row : header_row + 20]:
        for idx, column in enumerate(columns):
            value = _clean_cell(row[idx] if idx < len(row) else "")
            if value and value not in samples[column]:
                samples[column].append(value)
            samples[column] = samples[column][:3]
    return samples


def _inspection_notes(values: list[tuple[Any, ...]], header_row: int | None, columns: list[str]) -> list[str]:
    notes = []
    if header_row and header_row > 1:
        notes.append("Rows before detected header may contain titles, notes, or independent tables.")
    duplicate_headers = [header for header, count in Counter(columns).items() if count > 1]
    if duplicate_headers:
        notes.append(f"Duplicate-looking columns detected: {', '.join(duplicate_headers)}")
    blank_rows = sum(1 for row in values if not any(_clean_cell(value) for value in row))
    if blank_rows:
        notes.append(f"Blank rows detected: {blank_rows}")
    return notes


def _is_potential_lead(row: dict[str, Any]) -> bool:
    strong = any(row.get(field) for field in ["email", "phone", "website", "tax_id"])
    company_or_contact = any(row.get(field) for field in ["company_name", "contact_name"])
    return strong or company_or_contact


def _exact_duplicate_key(row: dict[str, Any]) -> str:
    for field in ["email", "phone", "tax_id", "website"]:
        if row.get(field):
            return f"{field}:{row[field]}"
    if row.get("company_name_normalized") and (row.get("city") or row.get("district")):
        return f"company_location:{row['company_name_normalized']}|{row.get('city') or row.get('district')}"
    return ""


def _possible_duplicate_key(row: dict[str, Any]) -> str:
    name = row.get("company_name_normalized", "")
    if not name:
        return ""
    token = "".join(name.split()[:2])
    return f"{token}|{row.get('city') or row.get('district') or _email_domain(row.get('email', ''))}"


def _merge_non_conflicting(canonical: dict[str, Any], row: dict[str, Any]) -> None:
    for field in STANDARD_FIELDS:
        if not canonical.get(field) and row.get(field):
            canonical[field] = row[field]
        elif canonical.get(field) and row.get(field) and canonical[field] != row[field] and field in {"secondary_email", "secondary_phone", "notes", "internal_notes"}:
            canonical[field] = _append_unique(canonical[field], row[field])


def _write_summary(wb: Workbook, prepared: PreparedData) -> None:
    ws = wb.create_sheet("SUMMARY")
    total_rows = sum(info.row_count for info in prepared.source_infos)
    missing_email = sum(1 for row in prepared.master_rows if not row.get("email"))
    invalid_email = sum(1 for row in prepared.master_rows if row.get("email_status") == "invalid")
    missing_phone = sum(1 for row in prepared.master_rows if not row.get("phone"))
    missing_website = sum(1 for row in prepared.master_rows if not row.get("website"))
    metrics = [
        ("Total source files", len({info.source_file for info in prepared.source_infos})),
        ("Total source sheets", len(prepared.source_infos)),
        ("Total rows inspected", total_rows),
        ("Total potential lead rows", len(prepared.source_rows)),
        ("Total valid standardized leads", len(prepared.master_rows)),
        ("Total exact duplicates", sum(1 for dup in prepared.duplicates if dup["duplicate_type"] == "exact")),
        ("Total possible duplicates", sum(1 for dup in prepared.duplicates if dup["duplicate_type"] == "possible")),
        ("Total excluded non-lead rows", len(prepared.excluded_rows)),
        ("Total rows requiring manual review", sum(1 for row in prepared.master_rows if row.get("review_required") == "true")),
        ("Missing email count", missing_email),
        ("Missing email percentage", _percentage(missing_email, len(prepared.master_rows))),
        ("Invalid email count", invalid_email),
        ("Invalid email percentage", _percentage(invalid_email, len(prepared.master_rows))),
        ("Missing phone count", missing_phone),
        ("Missing phone percentage", _percentage(missing_phone, len(prepared.master_rows))),
        ("Missing website count", missing_website),
        ("Missing website percentage", _percentage(missing_website, len(prepared.master_rows))),
        ("Data-quality summary", "Review rows marked review_required before outreach or enrichment."),
    ]
    ws.append(["metric", "value"])
    for metric in metrics:
        ws.append(list(metric))
    _format_sheet(ws)
    _add_table(ws, "SummaryTable")


def _write_source_index(wb: Workbook, prepared: PreparedData) -> None:
    headers = [
        "source_file",
        "source_sheet",
        "source_database",
        "original_row_count",
        "imported_lead_count",
        "duplicate_count",
        "error_count",
        "excluded_non_lead_count",
        "notes",
        "generated_source_tab_name",
    ]
    rows = [
        {
            "source_file": info.source_file,
            "source_sheet": info.source_sheet,
            "source_database": info.source_database,
            "original_row_count": info.row_count,
            "imported_lead_count": info.imported_lead_count,
            "duplicate_count": info.duplicate_count,
            "error_count": info.error_count,
            "excluded_non_lead_count": info.excluded_count,
            "notes": " | ".join(info.notes),
            "generated_source_tab_name": info.source_tab_name,
        }
        for info in prepared.source_infos
    ]
    _write_table_sheet(wb, "SOURCE_INDEX", headers, rows)


def _write_lookups(wb: Workbook, prepared: PreparedData) -> None:
    ws = wb.create_sheet("LOOKUPS")
    ws.append(["lookup_name", "value"])
    dynamic = {
        **LOOKUPS,
        "industries": sorted({row.get("industry", "") for row in prepared.master_rows if row.get("industry")}),
        "districts": sorted({row.get("district", "") for row in prepared.master_rows if row.get("district")}),
        "regions": sorted({row.get("region", "") for row in prepared.master_rows if row.get("region")}),
        "countries": sorted({row.get("country", "") for row in prepared.master_rows if row.get("country")}) or LOOKUPS["countries"],
    }
    for key, values in dynamic.items():
        for value in values:
            ws.append([key, value])
    _format_sheet(ws)
    _add_table(ws, "LookupsTable")


def _write_data_dictionary(wb: Workbook) -> None:
    headers = [
        "field_name",
        "description",
        "data_type",
        "required",
        "example",
        "normalization_rule",
        "LeadOps usage",
        "ForgeOS integration notes",
    ]
    rows = [definition.__dict__ for definition in FIELD_DEFINITIONS]
    _write_table_sheet(wb, "DATA_DICTIONARY", headers, rows)


def _write_table_sheet(wb: Workbook, sheet_name: str, headers: list[str], rows: Iterable[dict[str, Any]], highlight_review: bool = False) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    if not headers:
        headers = ["note"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    _format_sheet(ws, highlight_review)
    _add_table(ws, f"{re.sub(r'[^A-Za-z0-9]', '', sheet_name)[:20]}Table")


def _format_sheet(ws: Any, highlight_review: bool = False) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 55)
        ws.column_dimensions[column_cells[0].column_letter].width = max(width, 12)
    if highlight_review and ws.max_row > 1:
        headers = [cell.value for cell in ws[1]]
        if "review_required" in headers:
            review_col = headers.index("review_required") + 1
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=review_col).value).lower() == "true":
                    for cell in ws[row]:
                        cell.fill = review_fill


def _add_table(ws: Any, table_name: str) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    table = Table(displayName=table_name, ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def _flatten_source_row(row: dict[str, Any]) -> dict[str, Any]:
    standardized = row["standardized"]
    return {
        "source_file": standardized["source_file"],
        "source_sheet": standardized["source_sheet"],
        "source_row": standardized["source_row"],
        "source_database": standardized["source_database"],
        "standardized_lead_id": standardized["lead_id"],
        "duplicate_group_id": row.get("duplicate_group_id", ""),
        "duplicate_status": row.get("duplicate_status", standardized["duplicate_status"]),
        "validation_status": standardized["validation_status"],
        "review_required": standardized["review_required"],
        "raw_source_data": standardized["raw_source_data"],
    }


def _summary_payload(prepared: PreparedData, workbook_path: Path, report_path: Path, validation_path: Path, validation: dict[str, Any]) -> dict[str, Any]:
    return {
        "workbook_path": str(workbook_path),
        "inspection_report_path": str(report_path),
        "validation_report_path": str(validation_path),
        "source_files": sorted({info.source_file for info in prepared.source_infos}),
        "source_sheets": len(prepared.source_infos),
        "master_leads": len(prepared.master_rows),
        "exact_duplicates": sum(1 for dup in prepared.duplicates if dup["duplicate_type"] == "exact"),
        "possible_duplicates": sum(1 for dup in prepared.duplicates if dup["duplicate_type"] == "possible"),
        "import_errors": len(prepared.import_errors),
        "excluded_non_lead_rows": len(prepared.excluded_rows),
        "validation": validation,
    }


def _printable_result(result: dict[str, Any]) -> dict[str, Any]:
    prepared = result["prepared"]
    return {
        "workbook_path": str(result["workbook_path"]),
        "inspection_report_path": str(result["inspection_report_path"]),
        "validation_report_path": str(result["validation_report_path"]),
        "source_files_inspected": len({info.source_file for info in prepared.source_infos}),
        "source_sheets_inspected": len(prepared.source_infos),
        "master_leads": len(prepared.master_rows),
        "duplicates": len(prepared.duplicates),
        "import_errors": len(prepared.import_errors),
        "excluded_non_lead_rows": len(prepared.excluded_rows),
        "validation": result["validation"],
    }


def _duplicate_headers() -> list[str]:
    return [
        "duplicate_group_id",
        "duplicate_type",
        "confidence",
        "canonical_lead_id",
        "related_lead_id",
        "source_file",
        "source_row",
        "matching_fields",
        "conflicting_fields",
        "recommended_action",
        "review_status",
    ]


def _import_error_headers() -> list[str]:
    return [
        "source_file",
        "source_sheet",
        "source_row",
        "raw_source_data",
        "error_type",
        "error_description",
        "recommended_action",
    ]


def _validation_status(row: dict[str, Any]) -> str:
    if row.get("email_status") == "invalid" or row.get("phone_status") == "invalid" or row.get("website_status") == "invalid":
        return "review_required"
    if row.get("company_name") and (row.get("email") or row.get("phone") or row.get("website")):
        return "valid"
    return "partial"


def _data_quality_score(row: dict[str, Any]) -> int:
    score = 0
    for field in ["company_name", "email", "phone", "website", "city", "industry"]:
        if row.get(field):
            score += 15
    if row.get("validation_status") == "valid":
        score += 10
    return min(score, 100)


def _append_secondary(row: dict[str, Any], target: str, value: str) -> None:
    secondary = {
        "email": "secondary_email",
        "phone": "secondary_phone",
        "mobile_phone": "secondary_phone",
        "website": "other_social_url",
    }.get(target)
    if secondary:
        row[secondary] = _append_unique(row.get(secondary, ""), value)


def _append_unique(existing: str, value: str) -> str:
    if not value:
        return existing
    parts = [part.strip() for part in str(existing).split(";") if part.strip()]
    if value not in parts:
        parts.append(value)
    return "; ".join(parts)


def _split_multi_value(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[;,|\n]+", value) if part.strip()]


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _normalize_spaces(str(value))


def _normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _normalize_header(value: str) -> str:
    value = _normalize_spaces(value).casefold()
    value = re.sub(r"[_\-:/\\]+", " ", value)
    return _normalize_spaces(value)


def _source_database_label(path: Path) -> str:
    return path.parent.name


def _safe_sheet_suffix(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return cleaned or "Sheet"


def _stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}-{digest}"


def _email_domain(email: str) -> str:
    return email.split("@", 1)[1] if "@" in email else ""


def _percentage(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return "0%"
    return f"{numerator / denominator:.1%}"


def copy_template_only(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, destination)


if __name__ == "__main__":
    raise SystemExit(main())
