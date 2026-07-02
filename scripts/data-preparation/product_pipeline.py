"""Reusable product database preparation pipeline (staging-only, no ledger writes)."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from product_locale import (
    extract_packaging_quantity_hint,
    is_probable_ean,
    normalize_header,
    normalize_unit,
    parse_portuguese_date,
    parse_portuguese_decimal,
    today_iso,
)
from product_schema import (
    EXCLUDED_SOURCE_FIELDS,
    FIELD_ALIASES,
    FIELD_DEFINITIONS,
    SOURCE_COLUMNS,
    SOURCE_PRECEDENCE,
    STANDARD_FIELDS,
)

SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".tsv", ".html", ".htm"}
SCRIPT_VERSION = "2026.07.02"
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


@dataclass
class SourceInfo:
    source_id: str
    source_file: str
    source_path: str
    source_sheet: str
    source_database: str
    source_tab_name: str
    extension: str
    row_count: int
    column_count: int
    header_row: int | None
    columns: list[str]
    sample_values: dict[str, list[str]]
    merged_cells: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    imported_product_count: int = 0
    duplicate_count: int = 0
    error_count: int = 0
    excluded_count: int = 0
    excluded_field_hits: dict[str, int] = field(default_factory=dict)


@dataclass
class PreparedData:
    source_infos: list[SourceInfo]
    source_rows: list[dict[str, Any]]
    master_rows: list[dict[str, Any]]
    field_mappings: list[dict[str, Any]]
    duplicates: list[dict[str, Any]]
    import_errors: list[dict[str, Any]]
    quality_report: dict[str, Any]
    change_log: list[dict[str, Any]]


def discover_sources(input_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in input_dir.rglob("*")
        if path.is_file()
        and path.suffix.lower() in SUPPORTED_EXTENSIONS
        and not path.name.startswith("~$")
        and "standardized" not in path.name.lower()
        and "_files" not in path.parts
        and path.stat().st_size > 0
        and _looks_like_text_export(path)
    )


def _looks_like_text_export(path: Path) -> bool:
    sample = path.read_bytes()[:4096]
    if b"\x00" in sample and path.suffix.lower() in {".csv", ".tsv", ".html", ".htm"}:
        return False
    return True


def inspect_sources(input_dir: Path) -> list[SourceInfo]:
    source_infos: list[SourceInfo] = []
    for file_index, path in enumerate(discover_sources(input_dir), start=1):
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            source_infos.extend(_inspect_xlsx(path, file_index))
        elif suffix in {".html", ".htm"}:
            source_infos.append(_inspect_html(path, file_index))
        else:
            source_infos.append(_inspect_delimited(path, file_index))
    return _apply_source_precedence(source_infos)


def prepare_data(input_dir: Path, import_batch: str | None = None) -> PreparedData:
    import_batch = import_batch or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    source_infos = inspect_sources(input_dir)
    all_source_rows: list[dict[str, Any]] = []
    field_mappings: list[dict[str, Any]] = []
    import_errors: list[dict[str, Any]] = []
    change_log: list[dict[str, Any]] = []

    for info in source_infos:
        if info.notes and any("superseded" in note.lower() for note in info.notes):
            continue

        raw_rows = _read_source_rows(info)
        mappings = [_build_field_mapping(info, column, raw_rows) for column in info.columns]
        field_mappings.extend(mappings)
        mapping_by_column = {mapping["original_column"]: mapping["standardized_column"] for mapping in mappings}

        for raw_row in raw_rows:
            standardized, error = _standardize_row(raw_row, info, mapping_by_column, import_batch)
            if error:
                import_errors.append(error)
                info.error_count += 1
                continue

            if standardized.get("article_type", "").lower() == "portes de envio":
                info.excluded_count += 1
                continue

            info.imported_product_count += 1
            all_source_rows.append(
                {
                    **{key: standardized.get(key, "") for key in STANDARD_FIELDS if key != "raw_source_data"},
                    "raw_source_data": standardized.get("raw_source_data", ""),
                    "standardized": standardized,
                }
            )

    master_rows, duplicates = deduplicate_source_rows(all_source_rows)
    quality_report = build_quality_report(source_infos, master_rows, duplicates, import_errors, field_mappings)

    change_log.append(
        {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": "prepare_products",
            "script_version": SCRIPT_VERSION,
            "import_batch": import_batch,
            "source_count": len(source_infos),
            "master_count": len(master_rows),
            "duplicate_count": len(duplicates),
            "error_count": len(import_errors),
        }
    )

    return PreparedData(
        source_infos=source_infos,
        source_rows=all_source_rows,
        master_rows=master_rows,
        field_mappings=field_mappings,
        duplicates=duplicates,
        import_errors=import_errors,
        quality_report=quality_report,
        change_log=change_log,
    )


def deduplicate_source_rows(source_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    master_rows: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []
    canonical_by_code: dict[str, dict[str, Any]] = {}
    canonical_by_barcode: dict[str, dict[str, Any]] = {}
    possible_name_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for source_row in source_rows:
        row = dict(source_row["standardized"])
        code_key = (row.get("code") or "").strip().lower()
        barcode_key = (row.get("barcode") or "").strip()

        exact_key = ""
        if code_key and code_key in canonical_by_code:
            exact_key = f"code:{code_key}"
            canonical = canonical_by_code[code_key]
        elif barcode_key and barcode_key in canonical_by_barcode:
            exact_key = f"barcode:{barcode_key}"
            canonical = canonical_by_barcode[barcode_key]
        else:
            canonical = None

        if canonical is not None:
            duplicate_group_id = f"EXACT-{len(duplicates) + 1:05d}"
            row["duplicate_status"] = "exact_duplicate"
            row["review_required"] = "true"
            duplicates.append(
                {
                    "duplicate_group_id": duplicate_group_id,
                    "duplicate_type": "exact",
                    "confidence": "high",
                    "canonical_product_id": canonical["product_id"],
                    "related_product_id": row["product_id"],
                    "matching_fields": exact_key,
                    "recommended_action": "Keep canonical code; review price and designation drift.",
                    "review_status": "open",
                }
            )
            continue

        if code_key:
            canonical_by_code[code_key] = row
        if barcode_key:
            canonical_by_barcode[barcode_key] = row

        name_key = _possible_duplicate_key(row)
        if name_key:
            possible_name_groups[name_key].append(row)

        master_rows.append(row)

    for group_rows in possible_name_groups.values():
        if len(group_rows) < 2:
            continue
        duplicate_group_id = f"POSS-{len(duplicates) + 1:05d}"
        canonical = group_rows[0]
        for related in group_rows[1:]:
            if related["product_id"] == canonical["product_id"]:
                continue
            related["duplicate_status"] = "possible_duplicate"
            related["review_required"] = "true"
            duplicates.append(
                {
                    "duplicate_group_id": duplicate_group_id,
                    "duplicate_type": "possible",
                    "confidence": "medium",
                    "canonical_product_id": canonical["product_id"],
                    "related_product_id": related["product_id"],
                    "matching_fields": "normalized designation prefix",
                    "recommended_action": "Manual review before merge.",
                    "review_status": "open",
                }
            )

    for index, row in enumerate(master_rows, start=1):
        row["product_id"] = f"PROD-{index:06d}"

    return master_rows, duplicates


def build_quality_report(
    source_infos: list[SourceInfo],
    master_rows: list[dict[str, Any]],
    duplicates: list[dict[str, Any]],
    import_errors: list[dict[str, Any]],
    field_mappings: list[dict[str, Any]],
) -> dict[str, Any]:
    mapped_fields = Counter(mapping["standardized_column"] for mapping in field_mappings if mapping["standardized_column"])
    unmapped_columns = [mapping["original_column"] for mapping in field_mappings if not mapping["standardized_column"]]
    excluded_hits: Counter[str] = Counter()
    for info in source_infos:
        excluded_hits.update(info.excluded_field_hits)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "script_version": SCRIPT_VERSION,
        "source_files": len({info.source_file for info in source_infos}),
        "active_source_sheets": len([info for info in source_infos if not any("superseded" in n.lower() for n in info.notes)]),
        "total_source_rows": sum(info.row_count for info in source_infos),
        "master_product_rows": len(master_rows),
        "duplicate_groups": len(duplicates),
        "import_errors": len(import_errors),
        "rows_missing_barcode": sum(1 for row in master_rows if not row.get("barcode")),
        "rows_zero_purchase_price": sum(1 for row in master_rows if not row.get("purchase_price_ex_vat") or row.get("purchase_price_ex_vat") == "0"),
        "rows_zero_sale_price": sum(1 for row in master_rows if not row.get("sale_price_ex_vat") or row.get("sale_price_ex_vat") == "0"),
        "rows_with_packaging_hint": sum(1 for row in master_rows if row.get("packaging_quantity_hint")),
        "mapped_field_counts": dict(mapped_fields),
        "unmapped_columns_sample": sorted(set(unmapped_columns))[:20],
        "excluded_inventory_fields_detected": dict(excluded_hits),
        "inventory_quantities_reported": False,
    }


def build_workbook(prepared: PreparedData, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_summary(wb, prepared)
    _write_table_sheet(wb, "MASTER_PRODUCTS", STANDARD_FIELDS, prepared.master_rows, highlight_review=True)
    _write_table_sheet(wb, "FIELD_MAPPING", _field_mapping_columns(), prepared.field_mappings)
    _write_table_sheet(wb, "DUPLICATES", _duplicate_columns(), prepared.duplicates)
    _write_table_sheet(wb, "IMPORT_ERRORS", _error_columns(), prepared.import_errors)
    _write_table_sheet(wb, "SOURCE_PRECEDENCE", ["rank", "source_kind", "rule"], SOURCE_PRECEDENCE)
    _write_table_sheet(wb, "DATA_DICTIONARY", _dictionary_columns(), [_definition_row(defn) for defn in FIELD_DEFINITIONS])
    _write_quality_sheet(wb, prepared.quality_report)
    _write_change_log(wb, prepared.change_log)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def run_pipeline(input_dir: Path, output_dir: Path, workbook_name: str) -> dict[str, Any]:
    prepared = prepare_data(input_dir)
    workbook_path = output_dir / workbook_name
    build_workbook(prepared, workbook_path)
    report_path = output_dir / "products_quality_report.json"
    report_path.write_text(json.dumps(prepared.quality_report, indent=2), encoding="utf-8")
    write_inspection_report(prepared.source_infos, output_dir / "products_inspection_report.md")
    return {
        "workbook_path": str(workbook_path),
        "quality_report_path": str(report_path),
        "master_products": len(prepared.master_rows),
        "duplicates": len(prepared.duplicates),
        "errors": len(prepared.import_errors),
    }


def write_inspection_report(infos: list[SourceInfo], output_path: Path) -> None:
    lines = [
        "# Product source inspection (sanitized)",
        "",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
        "## Sources",
        "",
    ]
    for info in infos:
        lines.extend(
            [
                f"### {info.source_file} / {info.source_sheet}",
                f"- Extension: `{info.extension}`",
                f"- Rows: {info.row_count}",
                f"- Columns: {info.column_count}",
                f"- Active: {'no' if any('superseded' in n.lower() for n in info.notes) else 'yes'}",
                "",
            ]
        )
        if info.notes:
            lines.append("Notes:")
            lines.extend(f"- {note}" for note in info.notes)
            lines.append("")
        if info.excluded_field_hits:
            lines.append("Excluded inventory-related headers detected (counts only):")
            for key, count in info.excluded_field_hits.items():
                lines.append(f"- `{key}`: {count}")
            lines.append("")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")


def _apply_source_precedence(infos: list[SourceInfo]) -> list[SourceInfo]:
    artigos_csv = [
        info
        for info in infos
        if info.extension == ".csv" and "artigos" in info.source_file.lower()
    ]
    if artigos_csv:
        for info in infos:
            if info not in artigos_csv:
                info.notes.append("Superseded by native artigos CSV export per source precedence rank 2.")
        return infos

    csv_exports = [info for info in infos if info.extension == ".csv"]
    html_exports = [info for info in infos if info.extension in {".html", ".htm"}]
    if csv_exports and html_exports and len(csv_exports) == 1 and len(html_exports) == 1:
        csv_rows = csv_exports[0].row_count
        html_rows = html_exports[0].row_count
        if abs(csv_rows - html_rows) <= 1:
            html_exports[0].notes.append("Superseded by native CSV export per source precedence rank 2.")
    return infos


def _inspect_delimited(path: Path, file_index: int) -> SourceInfo:
    rows, encoding, delimiter = _read_delimited_file(path)
    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    samples = _sample_columns(headers, data_rows)
    excluded_hits = _count_excluded_headers(headers)
    return SourceInfo(
        source_id=f"SRC-{file_index:03d}",
        source_file=path.name,
        source_path=str(path),
        source_sheet=path.stem,
        source_database=path.parent.name,
        source_tab_name=path.stem,
        extension=path.suffix.lower(),
        row_count=len(data_rows),
        column_count=len(headers),
        header_row=1 if headers else None,
        columns=headers,
        sample_values=samples,
        notes=[f"Detected encoding: {encoding}", f"Delimiter: {repr(delimiter)}"],
        excluded_field_hits=excluded_hits,
    )


def _inspect_html(path: Path, file_index: int) -> SourceInfo:
    rows = _read_html_table(path)
    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    samples = _sample_columns(headers, data_rows)
    notes = ["HTML table export detected."]
    if headers and any(len(header) <= 4 for header in headers):
        notes.append("Headers appear truncated; prefer native CSV export.")
    return SourceInfo(
        source_id=f"SRC-{file_index:03d}",
        source_file=path.name,
        source_path=str(path),
        source_sheet=path.stem,
        source_database=path.parent.name,
        source_tab_name=path.stem,
        extension=path.suffix.lower(),
        row_count=len(data_rows),
        column_count=len(headers),
        header_row=1 if headers else None,
        columns=headers,
        sample_values=samples,
        notes=notes,
        excluded_field_hits=_count_excluded_headers(headers),
    )


def _inspect_xlsx(path: Path, file_index: int) -> list[SourceInfo]:
    wb = load_workbook(path, read_only=True, data_only=True)
    infos: list[SourceInfo] = []
    for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in (rows[0] if rows else [])]
        data_rows = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in rows[1:]
            if any(cell not in (None, "") for cell in row)
        ]
        infos.append(
            SourceInfo(
                source_id=f"SRC-{file_index:03d}-{sheet_index:02d}",
                source_file=path.name,
                source_path=str(path),
                source_sheet=sheet_name,
                source_database=path.parent.name,
                source_tab_name=sheet_name,
                extension=path.suffix.lower(),
                row_count=len(data_rows),
                column_count=len(headers),
                header_row=1 if headers else None,
                columns=headers,
                sample_values=_sample_columns(headers, data_rows),
                notes=[],
                excluded_field_hits=_count_excluded_headers(headers),
            )
        )
    wb.close()
    return infos


def _read_source_rows(info: SourceInfo) -> list[dict[str, Any]]:
    path = Path(info.source_path)
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[info.source_sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    elif path.suffix.lower() in {".html", ".htm"}:
        rows = _read_html_table(path)
    else:
        rows, _, _ = _read_delimited_file(path)

    headers = [str(value).strip() for value in (rows[0] if rows else [])]
    raw_rows: list[dict[str, Any]] = []
    for row_index, values in enumerate(rows[1:], start=2):
        if not any(str(value).strip() for value in values):
            continue
        record = {
            headers[idx]: str(values[idx]).strip() if idx < len(values) and values[idx] is not None else ""
            for idx in range(len(headers))
        }
        record["__source_row"] = row_index
        raw_rows.append(record)
    return raw_rows


def _standardize_row(
    raw_row: dict[str, Any],
    info: SourceInfo,
    mapping_by_column: dict[str, str],
    import_batch: str,
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    standardized: dict[str, Any] = {
        "source_file": info.source_file,
        "source_sheet": info.source_sheet,
        "source_row": raw_row.get("__source_row", ""),
        "source_database": info.source_database,
        "import_batch": import_batch,
        "validation_status": "valid",
        "duplicate_status": "unique",
        "review_required": "false",
        "data_quality_score": "100",
        "created_at": today_iso(),
        "updated_at": today_iso(),
        "raw_source_data": json.dumps({k: v for k, v in raw_row.items() if not k.startswith("__")}, ensure_ascii=False),
    }

    for column, value in raw_row.items():
        if column.startswith("__"):
            continue
        target = mapping_by_column.get(column, "")
        if not target or target in EXCLUDED_SOURCE_FIELDS:
            continue
        standardized[target] = value

    code = str(standardized.get("code", "")).strip()
    if not code:
        return standardized, {
            "source_file": info.source_file,
            "source_sheet": info.source_sheet,
            "source_row": raw_row.get("__source_row", ""),
            "error_type": "missing_code",
            "error_description": "Product code is required.",
            "recommended_action": "Fill reference/code or exclude row.",
        }

    designation = str(standardized.get("designation", "")).strip()
    if not designation:
        standardized["review_required"] = "true"
        standardized["validation_status"] = "review"

    barcode = re.sub(r"\D", "", str(standardized.get("barcode", "")).strip())
    if not barcode and is_probable_ean(code):
        barcode = re.sub(r"\D", "", code)
        standardized["review_required"] = "true"
        standardized["barcode"] = barcode

    for price_field in (
        "purchase_price_ex_vat",
        "purchase_price_inc_vat",
        "sale_price_ex_vat",
        "sale_price_inc_vat",
    ):
        normalized, status = parse_portuguese_decimal(standardized.get(price_field, ""))
        if status == "invalid":
            standardized["validation_status"] = "review"
            standardized["review_required"] = "true"
        standardized[price_field] = normalized

    pack_hint, pack_status = extract_packaging_quantity_hint(designation)
    if pack_status == "valid":
        standardized["packaging_quantity_hint"] = pack_hint

    standardized["sales_unit"] = normalize_unit(str(standardized.get("sales_unit", "")))
    standardized["capacity_unit"] = normalize_unit(str(standardized.get("capacity_unit", "")))

    if not standardized.get("sale_price_ex_vat"):
        standardized["review_required"] = "true"
        standardized["validation_status"] = "review"

    if not standardized.get("purchase_price_ex_vat"):
        standardized["review_required"] = "true"

    standardized["product_id"] = _stable_id("PROD", code)
    return standardized, None


def _build_field_mapping(info: SourceInfo, column: str, raw_rows: list[dict[str, Any]]) -> dict[str, Any]:
    normalized = normalize_header(column)
    standardized = FIELD_ALIASES.get(normalized, "")
    sample = next((row.get(column, "") for row in raw_rows if row.get(column)), "")
    confidence = "high" if standardized else "none"
    return {
        "source_file": info.source_file,
        "source_sheet": info.source_sheet,
        "original_column": column,
        "normalized_header": normalized,
        "standardized_column": standardized,
        "sample_value": _sanitize_sample(sample),
        "confidence": confidence,
        "import_allowed": "false" if standardized in EXCLUDED_SOURCE_FIELDS else "true",
    }


def _sanitize_sample(value: str) -> str:
    if not value:
        return ""
    if re.search(r"\d", value):
        return "<redacted-numeric-sample>"
    return value[:40] + ("…" if len(value) > 40 else "")


def _read_delimited_file(path: Path) -> tuple[list[list[str]], str, str]:
    raw_bytes = path.read_bytes()
    for encoding in CSV_ENCODINGS:
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    if text is None:
        text = raw_bytes.decode("latin-1", errors="replace")
        encoding = "latin-1"

    text = text.replace("\x00", "")
    delimiter = ";" if text.count(";") >= text.count(",") else ","
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    return rows, encoding, delimiter


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._cell_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self._current_row = []
        elif tag in {"td", "th"}:
            self._cell_parts = []

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._current_row is not None:
            self._current_row.append(" ".join(self._cell_parts).strip())
            self._cell_parts = []
        elif tag == "tr" and self._current_row is not None:
            if any(cell.strip() for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None

    def handle_data(self, data: str) -> None:
        if self._current_row is not None:
            self._cell_parts.append(data.strip())


def _read_html_table(path: Path) -> list[list[str]]:
    parser = _HtmlTableParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    return parser.rows


def _sample_columns(headers: list[str], rows: list[list[str]]) -> dict[str, list[str]]:
    samples: dict[str, list[str]] = {}
    for index, header in enumerate(headers):
        values: list[str] = []
        for row in rows:
            if index < len(row) and row[index]:
                values.append(_sanitize_sample(str(row[index])))
            if len(values) >= 3:
                break
        samples[header] = values
    return samples


def _count_excluded_headers(headers: list[str]) -> dict[str, int]:
    hits: Counter[str] = Counter()
    for header in headers:
        normalized = FIELD_ALIASES.get(normalize_header(header), normalize_header(header))
        if normalized in EXCLUDED_SOURCE_FIELDS:
            hits[normalized] += 1
    return dict(hits)


def _possible_duplicate_key(row: dict[str, Any]) -> str:
    designation = re.sub(r"\s+", " ", (row.get("designation") or "").strip().lower())
    if len(designation) < 12:
        return ""
    return designation[:24]


def _stable_id(prefix: str, seed: str) -> str:
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:10].upper()
    return f"{prefix}-{digest}"


def _field_mapping_columns() -> list[str]:
    return [
        "source_file",
        "source_sheet",
        "original_column",
        "normalized_header",
        "standardized_column",
        "sample_value",
        "confidence",
        "import_allowed",
    ]


def _duplicate_columns() -> list[str]:
    return [
        "duplicate_group_id",
        "duplicate_type",
        "confidence",
        "canonical_product_id",
        "related_product_id",
        "matching_fields",
        "recommended_action",
        "review_status",
    ]


def _error_columns() -> list[str]:
    return [
        "source_file",
        "source_sheet",
        "source_row",
        "error_type",
        "error_description",
        "recommended_action",
    ]


def _dictionary_columns() -> list[str]:
    return [
        "field_name",
        "description",
        "data_type",
        "required",
        "example",
        "normalization_rule",
        "forgeos_target",
    ]


def _definition_row(defn: Any) -> dict[str, str]:
    return {
        "field_name": defn.field_name,
        "description": defn.description,
        "data_type": defn.data_type,
        "required": "true" if defn.required else "false",
        "example": defn.example,
        "normalization_rule": defn.normalization_rule,
        "forgeos_target": defn.forgeos_target,
    }


def _write_summary(wb: Workbook, prepared: PreparedData) -> None:
    ws = wb.create_sheet("SUMMARY", 0)
    rows = [
        ["Product import staging workbook"],
        ["Generated at", datetime.now(timezone.utc).isoformat()],
        ["Script version", SCRIPT_VERSION],
        ["Master products", len(prepared.master_rows)],
        ["Duplicate groups", len(prepared.duplicates)],
        ["Import errors", len(prepared.import_errors)],
        ["Inventory quantities exported", "NO"],
    ]
    for row in rows:
        ws.append(row)


def _write_quality_sheet(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("QUALITY_REPORT")
    ws.append(["metric", "value"])
    for key, value in report.items():
        ws.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])


def _write_change_log(wb: Workbook, change_log: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("CHANGE_LOG")
    if not change_log:
        ws.append(["timestamp", "action"])
        return
    headers = list(change_log[0].keys())
    ws.append(headers)
    for entry in change_log:
        ws.append([entry.get(header, "") for header in headers])


def _xl_safe(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _write_table_sheet(
    wb: Workbook,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    highlight_review: bool = False,
) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    review_fill = PatternFill("solid", fgColor="FFF4CE")
    for row in rows:
        values = [_xl_safe(row.get(column, "")) for column in columns]
        ws.append(values)
        if highlight_review and str(row.get("review_required", "")).lower() == "true":
            for cell in ws[ws.max_row]:
                cell.fill = review_fill
    if rows:
        table = Table(displayName=title[:32].replace(" ", "_"), ref=f"A1:{_column_letter(len(columns))}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _column_letter(count: int) -> str:
    result = ""
    while count:
        count, remainder = divmod(count - 1, 26)
        result = chr(65 + remainder) + result
    return result
